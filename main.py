import os
import sys
import tempfile
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, Canvas, Frame, Label, Button, Scrollbar, Toplevel, ttk, StringVar, Scale
from pathlib import Path
from io import BytesIO
from PIL import Image as PILImage, ImageTk
import fitz  # PyMuPDF
from openpyxl import load_workbook
from docx2pdf import convert as docx2pdf_convert
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas as rl_canvas
import PyPDF2

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def convert_image_to_pdf(image_path):
    pdf_path = f"{os.path.splitext(image_path)[0]}.pdf"
    with PILImage.open(image_path) as img:
        img = img.convert("RGB")
        img.save(pdf_path, "PDF", resolution=100.0)
    return pdf_path

def convert_xlsx_to_pdf(xlsx_path):
    pdf_path = f"{os.path.splitext(xlsx_path)[0]}.pdf"
    workbook = load_workbook(xlsx_path)
    sheet = workbook.active

    c = rl_canvas.Canvas(pdf_path, pagesize=letter)
    width, height = letter

    y = height - 40
    for row in sheet.iter_rows(values_only=True):
        text = ", ".join([str(cell) for cell in row])
        c.drawString(40, y, text)
        y -= 20
        if y < 40:
            c.showPage()
            y = height - 40

    c.save()
    return pdf_path

def convert_docx_to_pdf(docx_path):
    pdf_path = f"{os.path.splitext(docx_path)[0]}.pdf"
    docx2pdf_convert(docx_path, pdf_path)
    return pdf_path

def convert_to_pdf(file_path):
    ext = file_path.lower().split('.')[-1]
    if ext in ['jpg', 'jpeg', 'png']:
        return convert_image_to_pdf(file_path)
    elif ext == 'xlsx':
        return convert_xlsx_to_pdf(file_path)
    elif ext == 'docx':
        return convert_docx_to_pdf(file_path)
    else:
        return file_path

def combine_files():
    root = tk.Tk()
    root.withdraw()

    file_paths = filedialog.askopenfilenames(
        title='Select files to combine',
        filetypes=(('Supported files', '*.pdf *.jpg *.jpeg *.png *.xlsx *.docx'), ('All files', '*.*')),
        initialdir=os.getcwd()
    )

    if not file_paths:
        messagebox.showerror("No File Selected", "Please select at least one file to combine.")
        return

    pdf_paths = []
    for file_path in file_paths:
        pdf_path = convert_to_pdf(file_path)
        if pdf_path:
            pdf_paths.append(pdf_path)

    def process_pdf_paths(pdf_paths):
        try:
            output_pdf = fitz.open()
            for pdf_path in pdf_paths:
                try:
                    input_pdf = fitz.open(pdf_path)
                    output_pdf.insert_pdf(input_pdf)
                    input_pdf.close()
                except Exception as e:
                    messagebox.showerror("PDF Processing Error", f"Failed to process {pdf_path}: {str(e)}")
                    continue

            default_output_path = Path.cwd() / "combined_output.pdf"
            output_pdf_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                initialdir=os.getcwd(),
                initialfile=default_output_path.name,
                title="Save Combined PDF"
            )

            if output_pdf_path:
                output_pdf.save(output_pdf_path)
                messagebox.showinfo("Success", "The files have been combined successfully.")
            else:
                messagebox.showerror("Save Error", "Failed to save the PDF. No file was chosen.")
        except Exception as e:
            messagebox.showerror("PDF Creation Error", f"Failed to create combined PDF: {str(e)}")
        finally:
            if 'output_pdf' in locals():
                output_pdf.close()

    threading.Thread(target=process_pdf_paths, args=(pdf_paths,)).start()

def rotate_pages():
    file_path = filedialog.askopenfilename(title="Select PDF", filetypes=[('PDF files', '*.pdf')])
    if not file_path:
        return

    pdf = fitz.open(file_path)
    rotation_states = [0] * len(pdf)

    rotate_input_window = Toplevel()
    rotate_input_window.title("Rotate and Sort Pages")
    rotate_input_window.geometry("800x600")

    canvas = Canvas(rotate_input_window, bg='white', borderwidth=0, highlightthickness=0)
    scrollbar = Scrollbar(rotate_input_window, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    canvas_frame = Frame(canvas, bg='white')
    canvas_window = canvas.create_window((0, 0), window=canvas_frame, anchor='nw', width=canvas.cget('width'))

    image_labels = []
    page_frames = []

    def rotate_page(page_num):
        rotation_states[page_num] = (rotation_states[page_num] + 90) % 360
        render_thumbnail(page_num)

    def render_thumbnail(page_num):
        page = pdf.load_page(page_num)
        rotation_degree = rotation_states[page_num]
        matrix = fitz.Matrix(0.25, 0.25).prerotate(rotation_degree)
        pix = page.get_pixmap(matrix=matrix)
        img = PILImage.open(BytesIO(pix.tobytes()))
        photo = ImageTk.PhotoImage(img)
        img_label, text_var = image_labels[page_num]
        img_label.config(image=photo)
        img_label.image = photo
        text_var.set(f"Page {page_num + 1} - {rotation_states[page_num]}°")

    def on_drag_start(event, frame, page_num):
        frame.drag_start_y = event.y_root
        frame.start_y = frame.winfo_y()

    def on_drag_motion(event, frame, page_num):
        dy = event.y_root - frame.drag_start_y
        frame.place(y=frame.start_y + dy)

    def on_drag_end(event, frame, page_num):
        new_y = canvas_frame.winfo_y() + frame.winfo_y() + event.y
        new_index = new_y // frame.winfo_height()
        old_index = page_frames.index(frame)

        frame_moved = page_frames.pop(old_index)
        page_frames.insert(new_index, frame_moved)

        for i, frame in enumerate(page_frames):
            frame.grid_forget()
            frame.grid(row=i, column=0, padx=10, pady=10, sticky='nsew')

        canvas_frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))

    for i in range(len(pdf)):
        frame = Frame(canvas_frame, relief=tk.SUNKEN, borderwidth=1, pady=5, bg='gray')
        frame.grid(row=i, column=0, padx=10, pady=10, sticky='nsew')
        page_frames.append(frame)

        text_var = StringVar(value=f"Page {i + 1} - 0°")
        label = Label(frame, textvariable=text_var)
        label.pack(side='top', pady=2)

        img_label = Label(frame)
        img_label.pack(side='top', fill=tk.BOTH, expand=True)
        image_labels.append((img_label, text_var))

        rotate_btn = Button(frame, text="Rotate", command=lambda i=i: rotate_page(i))
        rotate_btn.pack(side='bottom')

        frame.bind("<Button-1>", lambda e, frame=frame, page_num=i: on_drag_start(e, frame, page_num))
        frame.bind("<B1-Motion>", lambda e, frame=frame, page_num=i: on_drag_motion(e, frame, page_num))
        frame.bind("<ButtonRelease-1>", lambda e, frame=frame, page_num=i: on_drag_end(e, frame, page_num))

    def apply_rotation_and_save():
        sorted_pdf = fitz.open()

        for frame, (img_label, text_var) in zip(page_frames, image_labels):
            page_text = text_var.get()
            page_num = int(page_text.split()[1]) - 1
            rotation_degree = rotation_states[page_num]

            page = pdf.load_page(page_num)
            page.set_rotation(rotation_degree)

            # Create a new page with the rotated content
            new_page = sorted_pdf.new_page(width=page.rect.width, height=page.rect.height)
            new_page.show_pdf_page(new_page.rect, pdf, page_num, rotate=rotation_degree)

        save_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if save_path:
            sorted_pdf.save(save_path)
            messagebox.showinfo("Success", "PDF saved successfully with rotations and new order applied.")
            rotate_input_window.destroy()

    save_button = Button(rotate_input_window, text="Save Rotated PDF", command=apply_rotation_and_save)
    save_button.pack(side='bottom', pady=10)

    for i in range(len(pdf)):
        render_thumbnail(i)

    canvas_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

def split_pdf():
    root = tk.Tk()
    root.withdraw()

    input_pdf_path = filedialog.askopenfilename(title="Select PDF to Split", filetypes=[('PDF files', '*.pdf')])
    if not input_pdf_path:
        messagebox.showinfo("Info", "No file selected.")
        return

    try:
        pdf_reader = fitz.open(input_pdf_path)
        output_folder = filedialog.askdirectory(title="Select Folder to Save Split Pages")
        if not output_folder:
            return

        for page_num in range(len(pdf_reader)):
            pdf_writer = fitz.open()
            pdf_writer.insert_pdf(pdf_reader, from_page=page_num, to_page=page_num)

            output_path = os.path.join(output_folder, f"page_{page_num + 1}.pdf")
            pdf_writer.save(output_path)

        messagebox.showinfo("Success", "PDF split successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while splitting the PDF: {e}")
    finally:
        root.destroy()

def delete_pages():
    file_path = filedialog.askopenfilename(title="Select PDF", filetypes=[('PDF files', '*.pdf')])
    if not file_path:
        return

    pdf = fitz.open(file_path)
    num_pages = pdf.page_count

    delete_window = Toplevel()
    delete_window.title("Delete Pages")
    delete_window.geometry("800x600")

    canvas = Canvas(delete_window)
    scrollbar = Scrollbar(delete_window, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    pages_frame = Frame(canvas)
    canvas.create_window((0, 0), window=pages_frame, anchor="nw")

    def on_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    pages_frame.bind("<Configure>", on_configure)

    page_selections = []

    for i in range(num_pages):
        page_frame = Frame(pages_frame)
        page_frame.pack(fill="x", expand=True)

        chk_state = tk.BooleanVar()
        chk = tk.Checkbutton(page_frame, text=f"Page {i+1}", var=chk_state)
        chk.pack(side="left")
        page_selections.append(chk_state)

        page = pdf.load_page(i)
        pix = page.get_pixmap(dpi=72) 
        img_data = BytesIO(pix.tobytes("png"))
        img = PILImage.open(img_data)
        img.thumbnail((100, 100))
        photo = ImageTk.PhotoImage(img)
        
        img_label = Label(page_frame, image=photo)
        img_label.image = photo
        img_label.pack(side="left")

    def delete_selected_pages():
        pages_to_delete = [i for i, selected in enumerate(page_selections) if selected.get()]
        if not pages_to_delete:
            messagebox.showinfo("Info", "No pages selected for deletion.")
            return

        for page_num in sorted(pages_to_delete, reverse=True):
            pdf.delete_page(page_num)

        if len(pdf) == 0:
            messagebox.showerror("Error", "Cannot save a PDF with zero pages. Please ensure at least one page remains.")
            return

        save_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if save_path:
            pdf.save(save_path)
            messagebox.showinfo("Success", "Selected pages deleted successfully.")

        delete_window.destroy()

    delete_btn = Button(delete_window, text="Delete Selected Pages", command=delete_selected_pages)
    delete_btn.pack(pady=20)

    canvas.configure(scrollregion=canvas.bbox("all"))

    delete_window.mainloop()

def extract_images():
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(title="Select PDF", filetypes=[('PDF files', '*.pdf')])
    if not file_path:
        messagebox.showinfo("Info", "No PDF file selected.")
        root.destroy()
        return

    base_name = os.path.splitext(os.path.basename(file_path))[0]

    img_folder_path = filedialog.askdirectory(title="Select Folder to Save Images")
    if not img_folder_path:
        root.destroy()
        return

    def extract_images_from_pdf():
        try:
            pdf_file = fitz.open(file_path)
            for page_num, page in enumerate(pdf_file, start=1):
                image = page.get_pixmap(dpi=300)

                img_name = f"{base_name}_p{page_num:03}.png"
                img_path = os.path.join(img_folder_path, img_name)

                image.save(img_path)

            success_message = f"PDF extracted to {len(pdf_file)} images"
            messagebox.showinfo("Success", success_message)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to extract images from PDF: {e}")
        finally:
            pdf_file.close()
            root.destroy()

    threading.Thread(target=extract_images_from_pdf).start()

def convert_pdf_to_word():
    import pdf2docx

    root = tk.Tk()
    root.withdraw()

    pdf_file_path = filedialog.askopenfilename(title="Select a PDF file", filetypes=[("PDF files", "*.pdf")])
    if not pdf_file_path:
        messagebox.showerror("Error", "No PDF file selected.")
        root.destroy()
        return

    base_name = os.path.splitext(os.path.basename(pdf_file_path))[0]

    save_folder = filedialog.askdirectory(title="Select a folder to save the DOCX file")
    if not save_folder:
        root.destroy()
        return

    docx_file_path = os.path.join(save_folder, f"{base_name}_p.docx")

    def convert_pdf_to_docx():
        try:
            docx_converter = pdf2docx.Converter(pdf_file_path)
            docx_converter.convert(docx_file_path)
            docx_converter.close()
            messagebox.showinfo("Success", "PDF to DOCX conversion and saving completed successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to convert PDF file: {e}")
        finally:
            root.destroy()

    threading.Thread(target=convert_pdf_to_docx).start()

def convert_pdf_to_excel():
    import pdfplumber
    from openpyxl import Workbook

    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(title="Select PDF", filetypes=[('PDF files', '*.pdf')])
    if not file_path:
        messagebox.showinfo("Info", "No file selected.")
        return

    def convert_to_excel():
        try:
            with pdfplumber.open(file_path) as pdf:
                wb = Workbook()
                ws = wb.active

                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        for line in text.split('\n'):
                            ws.append([line])

                output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Save Excel File As")
                if output_path:
                    wb.save(output_path)
                    messagebox.showinfo("Success", "PDF converted to Excel successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
        finally:
            root.destroy()

    threading.Thread(target=convert_to_excel).start()

def show_slider_and_optimize(pdf_file_path):
    optimization_window = Toplevel()
    optimization_window.title("Optimization Level")
    optimization_window.configure(bg='#F0F0F0')
    optimization_window.geometry("400x400")

    title_label = Label(optimization_window, text="PDF Optimization", font=("Helvetica", 16, "bold"), bg='#F0F0F0', fg='#0000FF')
    title_label.pack(pady=10)

    optimization_slider = Scale(optimization_window, from_=0, to=100, orient=tk.HORIZONTAL, label="Optimization Level (%)", bg='#F0F0F0', troughcolor='#0000FF')
    optimization_slider.pack(pady=20, padx=20)

    size_label = Label(optimization_window, text="Estimated File Size: ", font=("Helvetica", 10), bg='#F0F0F0', fg='#222222')
    size_label.pack()

    def update_size_label(value):
        estimated_size = calculate_estimated_size(pdf_file_path, int(value))
        size_label.config(text=f"Estimated File Size: {estimated_size:.2f} KB")

    optimization_slider.bind("<Motion>", lambda event: update_size_label(optimization_slider.get()))
    optimization_slider.bind("<ButtonRelease-1>", lambda event: update_size_label(optimization_slider.get()))

    optimize_now_btn = Button(optimization_window, text="Optimize Now", command=lambda: optimize_pdf(pdf_file_path, optimization_slider.get(), optimization_window), bg="#4CAF50", fg="white", relief=tk.FLAT)
    optimize_now_btn.pack(pady=10)

    center_window_on_screen(optimization_window)


def center_window_on_screen(window):
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry(f'{width}x{height}+{x}+{y}')

def calculate_estimated_size(pdf_file_path, optimization_level):
    original_size = os.path.getsize(pdf_file_path) / 1024
    estimated_size = original_size * (1 - optimization_level / 100)
    return estimated_size

def optimize_pdf(pdf_file_path, optimization_level, optimization_window):
    pdf_writer = PyPDF2.PdfWriter()
    pdf_file = fitz.open(pdf_file_path)

    for page_num in range(len(pdf_file)):
        page = pdf_file[page_num]
        pixmap = page.get_pixmap()
        with tempfile.NamedTemporaryFile(delete=True, suffix=".png") as temp_file:
            pixmap.save(temp_file.name, "png")
            image = PILImage.open(temp_file.name)

            img_pdf_stream = tempfile.NamedTemporaryFile(delete=True, suffix=".pdf")
            image.save(img_pdf_stream, "PDF", resolution=optimization_level)
            img_pdf_stream.seek(0)

            img_pdf = PyPDF2.PdfReader(img_pdf_stream)
            pdf_writer.add_page(img_pdf.pages[0])

    output_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
    if not output_path:
        return

    with open(output_path, "wb") as out:
        pdf_writer.write(out)
    
    messagebox.showinfo("Success", "PDF optimized")
    optimization_window.destroy()


def on_optimize_button_click():
    file_path = filedialog.askopenfilename(title="Select PDF", filetypes=[('PDF files', '*.pdf')])
    if file_path:
        show_slider_and_optimize(file_path)

def perform_ocr_and_convert_to_word():
    import pdf2docx

    pdf_file_path = filedialog.askopenfilename(title="Select PDF for OCR", filetypes=[('PDF files', '*.pdf')])
    if not pdf_file_path:
        return

    encodings = ['utf-8', 'latin-1', 'cp1252']

    def ocr_and_convert():
        try:
            with fitz.open(pdf_file_path) as pdf_document:
                ocr_text = ""
                for page_num in range(len(pdf_document)):
                    page = pdf_document[page_num]
                    for encoding in encodings:
                        try:
                            ocr_text += page.get_text("text")
                            break
                        except UnicodeDecodeError:
                            pass

            save_directory = filedialog.askdirectory(title="Select a folder to save the Word document")
            if not save_directory:
                return

            pdf_file_name = os.path.splitext(os.path.basename(pdf_file_path))[0]
            html_file_path = os.path.join(save_directory, f"{pdf_file_name}.html")

            with open(html_file_path, 'w', encoding='utf-8') as html_file:
                html_file.write(ocr_text)

            docx_file_path = os.path.join(save_directory, f"{pdf_file_name}.docx")
            pdf2docx.convert_file(html_file_path, 'docx', outputfile=docx_file_path)

            os.remove(html_file_path)

            messagebox.showinfo("Success", "OCR and Word conversion completed successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to perform OCR and convert to Word: {e}")

    threading.Thread(target=ocr_and_convert).start()

def start_app():
    root = tk.Tk()
    style = ttk.Style()
    root.title("PDF Tools")
    root.geometry("600x600")

    style.configure('TButton', font=("Roboto", 10), padding=10, borderwidth=1, relief="solid")
    style.map('TButton',
              foreground=[('pressed', 'red'), ('active', 'blue')],
              background=[('pressed', '!disabled', 'black'), ('active', 'white')])
    style.configure('TLabel', font=("Roboto", 12), background="white", foreground="black")

    title_label = ttk.Label(root, text="PDF Tools", font=("Roboto", 18, "bold"))
    title_label.pack(pady=20)
    subtitle_label = ttk.Label(root, text="By Wajdy Alzayer", font=("Roboto", 10))
    subtitle_label.pack()

    def open_paypal(event):
        webbrowser.open_new("https://paypal.me/WAlzayer")
        
    paypal_link = ttk.Label(root, text="Donate", font=("Roboto", 10, "underline"), cursor="hand2")
    paypal_link.bind("<Button-1>", open_paypal)
    paypal_link.pack(pady=10)

    frame = ttk.Frame(root)
    frame.pack(pady=20, padx=20, fill='both', expand=True)

    function_mappings = [
        ("Combine PDFs", combine_files),
        ("Split PDF", split_pdf),
        ("PDF to Images", extract_images),
        ("Convert to Word", convert_pdf_to_word),
        ("Convert to Excel", convert_pdf_to_excel),
        ("Delete Pages", delete_pages),
        ("Optimize PDF", on_optimize_button_click),
        ("Rotate Pages", rotate_pages),
        ("OCR (TBA)", perform_ocr_and_convert_to_word)
    ]

    icons = ["combine.png", "split.png", "toimages.png", "toword.png", "toexcel.png", "delete.png", "optimize.png", "rotate.png", "ocr.png"]

    for index, (label, command) in enumerate(function_mappings):
        row, col = divmod(index, 3)
        icon_path = resource_path(os.path.join("icons", icons[index]))
        icon_image = PILImage.open(icon_path).resize((50, 50), PILImage.Resampling.LANCZOS)
        photo = ImageTk.PhotoImage(icon_image)
        btn = ttk.Button(frame, text=label, image=photo, compound="top", command=command)
        btn.image = photo
        btn.grid(row=row, column=col, padx=10, pady=10, sticky='ewns')

    for i in range(3):
        frame.columnconfigure(i, weight=1)
    for i in range((len(function_mappings) + 2) // 3):
        frame.rowconfigure(i, weight=1)

    root.mainloop()

if __name__ == "__main__":
    start_app()
